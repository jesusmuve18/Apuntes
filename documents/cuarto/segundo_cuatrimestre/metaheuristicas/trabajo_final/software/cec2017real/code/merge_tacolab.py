"""
merge_tacolab.py — Fusiona los Excel de varios algoritmos en uno
               compatible con TacoLab.

Uso (desde code/):
    python3 merge_tacolab.py 10
    python3 merge_tacolab.py 30
    python3 merge_tacolab.py 50

Genera: tacolab_<dim>.xlsx listo para subir a tacolab.org/bench
como fichero único con los tres algoritmos.

Estructura de salida:
    alg      | milestone | F01 | F02 | ... | F30
    SCA      |     1     | ... | ... | ... | ...
    SCA      |     2     | ... | ... | ... | ...
    ...
    SCA_LS   |     1     | ... | ... | ... | ...
    ...
    ASCA     |     1     | ... | ... | ... | ...
    ...
"""

import sys
import openpyxl

def merge(dim):
    algoritmos = [
        ("SCA",    f"results_sca/results_cec2017_{dim}.xlsx"),
        ("SCA_LS", f"results_sca_ls/results_cec2017_{dim}.xlsx"),
        ("ASCA",   f"results_asca/results_cec2017_{dim}.xlsx"),
    ]

    output = f"tacolab_{dim}.xlsx"
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Sheet1"

    header_written = False

    for alg_name, filepath in algoritmos:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]   # (None, 'milestone', 'F01', ...)
        data   = rows[1:]  # filas de resultados

        # La cabecera original tiene None en col 0 (índice numérico del alg).
        # La sustituimos por 'alg', que es el nombre que espera TacoLab.
        if not header_written:
            ws_out.append(["alg"] + list(header[1:]))
            header_written = True

        for row in data:
            # Col 0 = nombre del algoritmo, resto igual que el original
            ws_out.append([alg_name] + list(row[1:]))

    wb_out.save(output)
    print(f"[OK] {output} generado con {len(algoritmos)} algoritmos.")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python3 merge_tacolab.py <dim>  (ej: 10, 30, 50)")
        sys.exit(1)
    merge(sys.argv[1])